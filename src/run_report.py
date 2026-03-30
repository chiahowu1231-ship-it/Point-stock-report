# src/run_report.py
# 券商分點狙擊分析 — v11 升級版
# 修正: 淨賣超也顯示均價/乖離率
# 新增: top_preview、大盤籌碼、AI 分析整合

import os
import re
import json
import time
import random
from datetime import datetime
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from bs4 import BeautifulSoup
import yfinance as yf

from config import LEGEND_BROKERS

# 大盤籌碼模組
try:
    from market_data import fetch_all_market_data
    HAS_MARKET_DATA = True
except ImportError:
    HAS_MARKET_DATA = False
    print("[WARN] market_data 模組未找到，跳過大盤籌碼抓取")

TZ = ZoneInfo("Asia/Taipei")
DEFAULT_DAYS = 5
HEAD_BROKER = "9600"
BASE_URL = "https://fubon-ebrokerdj.fbs.com.tw/z/zg/zgb/zgb0.djhtm"

UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

PRICE_CACHE = {}


def ensure_output_dir():
    os.makedirs("output", exist_ok=True)


def now_taipei():
    return datetime.now(TZ)


def safe_int(s: str) -> int:
    if s is None:
        return 0
    s = s.strip().replace(",", "")
    if s in ("", "--", "-"):
        return 0
    try:
        return int(s)
    except ValueError:
        m = re.search(r"-?\d+", s)
        return int(m.group()) if m else 0


def fetch_html(url: str, timeout: int = 20, retries: int = 5, base_sleep: float = 1.0) -> str:
    headers = {
        "User-Agent": UA,
        "Referer": "https://fubon-ebrokerdj.fbs.com.tw/",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "zh-TW,zh;q=0.9,en;q=0.8",
    }
    last_err = None
    for attempt in range(retries):
        try:
            r = requests.get(url, headers=headers, timeout=timeout)
            for enc in ("cp950", "big5", "utf-8"):
                try:
                    r.encoding = enc
                    if r.status_code == 200 and r.text and len(r.text) > 100:
                        return r.text
                except Exception:
                    continue
            last_err = RuntimeError(f"HTTP {r.status_code}")
        except Exception as e:
            last_err = e
        time.sleep(base_sleep * (2 ** attempt) + random.uniform(0, 0.5))
    raise last_err


def parse_table(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    res = {}
    pat = re.compile(
        r"GenLink2stk\(\s*['\"]AS(\w+)['\"]\s*,\s*['\"](.+?)['\"]\s*\)"
    )
    for tr in soup.find_all("tr"):
        tr_text = tr.decode_contents()
        m = pat.search(tr_text)
        if not m:
            continue
        sid = m.group(1)
        name = m.group(2)
        tds = tr.find_all("td", class_=re.compile(r"t3n[01]"))
        if len(tds) < 3:
            tds = tr.find_all("td")
            if len(tds) < 3:
                continue
        buy_txt = tds[0].get_text(strip=True)
        sell_txt = tds[1].get_text(strip=True)
        net_txt = tds[2].get_text(strip=True)
        res[sid] = {
            "name": name,
            "buy": buy_txt,
            "sell": sell_txt,
            "net": safe_int(net_txt),
        }
    return res


def get_stock_price(sid: str) -> float:
    if sid in PRICE_CACHE:
        return PRICE_CACHE[sid]
    price = 0.0
    for suffix in [".TW", ".TWO"]:
        try:
            data = yf.Ticker(f"{sid}{suffix}").history(period="5d")
            if not data.empty:
                price = float(data["Close"].iloc[-1])
                if price > 0:
                    break
        except Exception:
            continue
    PRICE_CACHE[sid] = price
    return price


def build_report(days: int):
    start_time = now_taipei()
    rows = []
    failures = []
    errors = []
    broker_ok = 0
    broker_fail = 0

    for broker_id, broker_name in LEGEND_BROKERS.items():
        url_qty = f"{BASE_URL}?a={HEAD_BROKER}&b={broker_id}&c=E&d={days}"
        url_amt = f"{BASE_URL}?a={HEAD_BROKER}&b={broker_id}&c=B&d={days}"

        try:
            html_qty = fetch_html(url_qty)
            html_amt = fetch_html(url_amt)
            qty_map = parse_table(html_qty)
            amt_map = parse_table(html_amt)

            for sid, info in qty_map.items():
                if sid not in amt_map:
                    continue

                net_qty = info["net"]
                net_amt = amt_map[sid]["net"]  # 仟元

                # ✅ 修正：淨賣超也計算均價
                # 淨買超 → 買入均價 | 淨賣超 → 賣出均價（取絕對值）
                if net_qty != 0 and net_amt != 0:
                    avg_cost = round(abs(net_amt) / abs(net_qty), 2)
                else:
                    avg_cost = 0.0

                price = get_stock_price(sid)

                if price > 0 and avg_cost > 0:
                    bias = round(((price - avg_cost) / avg_cost) * 100, 2)
                else:
                    bias = 0.0

                rows.append({
                    "日期": start_time.strftime("%Y%m%d"),
                    "代碼": sid,
                    "名稱": info["name"],
                    "大戶": broker_name,
                    "買進": info["buy"],
                    "賣出": info["sell"],
                    "淨超": net_qty,
                    "區間均價": avg_cost,
                    "現價": round(price, 2),
                    "乖離率": f"{bias}%",
                })

            broker_ok += 1
            time.sleep(1.5 + random.uniform(0, 0.5))

        except Exception as e:
            broker_fail += 1
            msg = f"{broker_name}({broker_id}) 失敗：{str(e)}"
            errors.append(msg)
            failures.append({
                "日期": start_time.strftime("%Y%m%d"),
                "分點代號": broker_id,
                "分點名稱": broker_name,
                "錯誤訊息": str(e),
                "網址(張數E)": url_qty,
                "網址(金額B)": url_amt,
            })

    df = pd.DataFrame(rows, columns=[
        "日期", "代碼", "名稱", "大戶", "買進", "賣出", "淨超", "區間均價", "現價", "乖離率"
    ])
    fail_df = pd.DataFrame(failures, columns=[
        "日期", "分點代號", "分點名稱", "錯誤訊息", "網址(張數E)", "網址(金額B)"
    ])

    summary = {
        "generated_at": start_time.isoformat(),
        "timezone": "Asia/Taipei",
        "days": days,
        "total_rows": int(len(df)),
        "brokers_total": int(len(LEGEND_BROKERS)),
        "brokers_ok": broker_ok,
        "brokers_fail": broker_fail,
        "success": (broker_fail == 0),
        "errors": errors[:50],
    }

    # ===== 排序 + top_preview =====
    TOP_N = int(os.getenv("TOP_N", "10"))

    if not df.empty:
        df["淨超"] = pd.to_numeric(df["淨超"], errors="coerce").fillna(0).astype(int)
        broker_order = (
            df.groupby("大戶")["淨超"].sum()
              .sort_values(ascending=False).index.tolist()
        )
        df["大戶"] = pd.Categorical(df["大戶"], categories=broker_order, ordered=True)
        df = df.sort_values(["大戶", "淨超"], ascending=[True, False]).reset_index(drop=True)

    top_preview = []
    if not df.empty and isinstance(df.get("大戶").dtype, pd.CategoricalDtype):
        for broker in df["大戶"].cat.categories.tolist():
            sub = df[df["大戶"] == broker].head(TOP_N)
            if sub.empty:
                continue
            top_preview.append({
                "broker": str(broker),
                "total_net": int(df[df["大戶"] == broker]["淨超"].sum()),
                "rows": [
                    {
                        "sid": r["代碼"], "name": r["名稱"],
                        "net": int(r["淨超"]), "avg": r["區間均價"],
                        "price": r["現價"], "bias": r["乖離率"],
                    }
                    for _, r in sub.iterrows()
                ]
            })

    summary["top_preview"] = top_preview
    summary["top_n"] = TOP_N

    if summary["total_rows"] == 0:
        summary["success"] = False
        summary["errors"] = (summary.get("errors") or []) + ["總資料筆數為 0"]

    return df, fail_df, summary


def export_excel(df, fail_df, xlsx_path):
    from openpyxl.styles import PatternFill, Font
    from openpyxl.formatting.rule import FormulaRule

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Report")
        ws = writer.sheets["Report"]
        for col, w in {"A": 10, "B": 8, "C": 14, "D": 28, "E": 10, "F": 10, "G": 10, "H": 12, "I": 10, "J": 10}.items():
            ws.column_dimensions[col].width = w

        last_row = ws.max_row
        if last_row >= 2:
            rng = f"J2:J{last_row}"
            fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fill_blue = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            fill_red = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
            font_white = Font(color="FFFFFF", bold=True)
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[r'=VALUE(SUBSTITUTE($J2,"%",""))>10'], fill=fill_red, font=font_white, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[r'=AND(VALUE(SUBSTITUTE($J2,"%",""))>3,VALUE(SUBSTITUTE($J2,"%",""))<=10)'], fill=fill_blue, stopIfTrue=True))
            ws.conditional_formatting.add(rng, FormulaRule(
                formula=[r'=VALUE(SUBSTITUTE($J2,"%",""))<=3'], fill=fill_green, stopIfTrue=True))

        # TopByBroker sheet
        top_n = int(os.getenv("TOP_N", "10"))
        if df is not None and not df.empty:
            top_df = df.groupby("大戶", sort=False).head(top_n)
            top_df.to_excel(writer, index=False, sheet_name="TopByBroker")

        if fail_df is not None and not fail_df.empty:
            fail_df.to_excel(writer, index=False, sheet_name="Failures")


def export_pdf(df, pdf_path, summary):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_path = os.path.join("fonts", "NotoSansTC-Regular.ttf")
    font_name = "NotoSansTC"
    use_cjk = False
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
            use_cjk = True
        except Exception:
            pass
    if not use_cjk:
        font_name = "Helvetica"

    doc = SimpleDocTemplate(pdf_path, pagesize=landscape(A4),
                            leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name
    styles["Normal"].fontName = font_name

    title = Paragraph("【券商分點狙擊分析】報表", styles["Title"])
    meta = Paragraph(
        f"產生時間：{summary['generated_at']}（{summary['timezone']}）　"
        f"查詢區間：{summary['days']} 日　"
        f"結果：{'成功' if summary['success'] else '部分失敗'}　"
        f"筆數：{summary['total_rows']}", styles["Normal"])
    elements = [title, Spacer(1, 8), meta, Spacer(1, 12)]

    header = list(df.columns) if df is not None and not df.empty else [
        "日期", "代碼", "名稱", "大戶", "買進", "賣出", "淨超", "區間均價", "現價", "乖離率"]
    data = [header] + (df.astype(str).values.tolist() if df is not None and not df.empty else [])
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), font_name), ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EAEAEA")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F7F7F7")]),
    ]))
    elements.append(table)
    doc.build(elements)


def main():
    ensure_output_dir()
    days = int(os.getenv("DAYS", str(DEFAULT_DAYS)))

    df, fail_df, summary = build_report(days)

    ymd = now_taipei().strftime("%Y%m%d")
    xlsx_path = os.path.join("output", f"IKE_Report_{ymd}.xlsx")
    pdf_path = os.path.join("output", f"IKE_Report_{ymd}.pdf")
    summary_path = os.path.join("output", "summary.json")

    export_excel(df, fail_df, xlsx_path)

    try:
        export_pdf(df, pdf_path, summary)
    except Exception as e:
        print(f"[WARN] PDF 輸出失敗（非致命）: {e}")

    # ===== 大盤籌碼資料 =====
    if HAS_MARKET_DATA and os.getenv("MARKET_DATA", "1") == "1":
        try:
            top_stock_ids = []
            seen = set()
            for block in (summary.get("top_preview") or [])[:5]:
                for r in (block.get("rows") or [])[:3]:
                    sid = r.get("sid", "")
                    if sid and sid not in seen:
                        top_stock_ids.append(sid)
                        seen.add(sid)

            market = fetch_all_market_data(
                top_stock_ids=top_stock_ids[:10],
                history_days=int(os.getenv("MARKET_HISTORY_DAYS", "6")),
            )
            summary["market_data"] = market
            print(f"[OK] 大盤籌碼: inst={len(market.get('institutional', []))}")
        except Exception as e:
            print(f"[WARN] 大盤籌碼失敗（非致命）: {e}")
            summary["market_data"] = {"fetch_errors": [str(e)]}

    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    print(f"[OK] Excel: {xlsx_path}")
    print(f"[OK] PDF  : {pdf_path}")
    print(f"[OK] Summary: {summary_path}")

    if summary.get("total_rows", 0) == 0 or summary.get("brokers_ok", 0) == 0:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
